import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QDoubleSpinBox,
)


Row = Dict[str, Any]


@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    excel_header: str
    kind: str  # "text" | "number"
    computed: bool = False
    default: Optional[float] = None
    decimals: int = 2
    min_value: float = 0.0
    max_value: float = 1_000_000_000.0
    number_format: Optional[str] = None  # openpyxl format, e.g. "0.00"


def compute_retail(row: Row) -> float:
    purchase = float(row.get("purchase_price") or 0.0)
    markup = float(row.get("markup_percent") or 0.0)
    return round(purchase * (1.0 + markup / 100.0), 2)


FIELDS: List[FieldSpec] = [
    FieldSpec(
        key="name",
        label="Наименование товара",
        excel_header="Наименование",
        kind="text",
        computed=False,
    ),
    FieldSpec(
        key="purchase_price",
        label="Закупочная цена",
        excel_header="Закупочная цена",
        kind="number",
        computed=False,
        default=0.0,
        decimals=2,
        min_value=0.0,
        max_value=1_000_000_000.0,
        number_format="0.00",
    ),
    FieldSpec(
        key="markup_percent",
        label="Наценка %",
        excel_header="Процент наценки",
        kind="number",
        computed=False,
        default=60.0,
        decimals=2,
        min_value=0.0,
        max_value=10_000.0,
        number_format="0.00",
    ),
    FieldSpec(
        key="retail_price",
        label="Розничная цена",
        excel_header="Розничная цена",
        kind="number",
        computed=True,
        decimals=2,
        min_value=0.0,
        max_value=1_000_000_000.0,
        number_format="0.00",
    ),
]

COMPUTED: Dict[str, Callable[[Row], Any]] = {"retail_price": compute_retail}


class PriceListApp(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Прайс-лист (PyQt6 + Excel)")
        self.resize(820, 520)

        self.rows: List[Row] = []
        self.input_widgets: Dict[str, Any] = {}
        self.computed_labels: Dict[str, QLabel] = {}

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        root.addWidget(self._build_inputs_group())
        root.addWidget(self._build_table())
        root.addLayout(self._build_actions())

        self._recompute_computed_fields()

    def _build_inputs_group(self) -> QGroupBox:
        box = QGroupBox("Ввод товара")
        layout = QFormLayout(box)
        layout.setLabelAlignment(Qt.AlignmentFlag.AlignLeft)
        layout.setFormAlignment(Qt.AlignmentFlag.AlignTop)
        layout.setHorizontalSpacing(12)
        layout.setVerticalSpacing(8)

        for spec in FIELDS:
            if spec.computed:
                value = QLabel("0.00")
                value.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                value.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
                self.computed_labels[spec.key] = value
                layout.addRow(spec.label, value)
                continue

            widget = self._create_input_widget(spec)
            self.input_widgets[spec.key] = widget
            layout.addRow(spec.label, widget)

        return box

    def _create_input_widget(self, spec: FieldSpec):
        if spec.kind == "text":
            w = QLineEdit()
            w.setPlaceholderText("Введите текст…")
            w.textChanged.connect(self._recompute_computed_fields)
            return w

        w = QDoubleSpinBox()
        w.setDecimals(spec.decimals)
        w.setRange(spec.min_value, spec.max_value)
        w.setSingleStep(1.0)
        if spec.default is not None:
            w.setValue(float(spec.default))
        w.valueChanged.connect(self._recompute_computed_fields)
        return w

    def _build_table(self) -> QTableWidget:
        table = QTableWidget(0, len(FIELDS))
        table.setHorizontalHeaderLabels([f.excel_header for f in FIELDS])
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setAlternatingRowColors(True)
        table.horizontalHeader().setStretchLastSection(True)
        self.table = table
        return table

    def _build_actions(self) -> QHBoxLayout:
        row = QHBoxLayout()
        row.setSpacing(10)

        self.add_btn = QPushButton("Добавить в таблицу")
        self.add_btn.clicked.connect(self.add_row)
        row.addWidget(self.add_btn)

        self.import_btn = QPushButton("Загрузить из Excel")
        self.import_btn.clicked.connect(self.import_from_excel)
        row.addWidget(self.import_btn)

        self.clear_btn = QPushButton("Очистить список")
        self.clear_btn.clicked.connect(self.clear_rows)
        row.addWidget(self.clear_btn)

        self.delete_btn = QPushButton("Удалить выбранное")
        self.delete_btn.clicked.connect(self.delete_selected_rows)
        row.addWidget(self.delete_btn)

        row.addStretch(1)

        self.export_btn = QPushButton("Сохранить отчет в Excel")
        self.export_btn.clicked.connect(self.export_excel)
        row.addWidget(self.export_btn)

        return row

    def _get_current_input_row(self) -> Row:
        row: Row = {}
        for spec in FIELDS:
            if spec.computed:
                continue
            w = self.input_widgets[spec.key]
            if spec.kind == "text":
                row[spec.key] = w.text().strip()
            else:
                row[spec.key] = float(w.value())

        for spec in FIELDS:
            if spec.computed:
                row[spec.key] = COMPUTED[spec.key](row)
        return row

    def _recompute_computed_fields(self) -> None:
        row = self._get_current_input_row()
        for spec in FIELDS:
            if not spec.computed:
                continue
            value = row.get(spec.key)
            if isinstance(value, (int, float)):
                self.computed_labels[spec.key].setText(f"{value:.{spec.decimals}f}")
            else:
                self.computed_labels[spec.key].setText(str(value))

    def add_row(self) -> None:
        row = self._get_current_input_row()
        if not row.get("name"):
            QMessageBox.warning(self, "Проверка", "Введите наименование товара.")
            return

        self.rows.append(row)
        self._append_row_to_table(row)
        self._reset_inputs_after_add()

    def _append_row_to_table(self, row: Row) -> None:
        r = self.table.rowCount()
        self.table.insertRow(r)
        for c, spec in enumerate(FIELDS):
            value = row.get(spec.key, "")
            if isinstance(value, (int, float)):
                text = f"{float(value):.{spec.decimals}f}"
            else:
                text = str(value)
            item = QTableWidgetItem(text)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(r, c, item)

        self.table.resizeColumnsToContents()

    def _reset_inputs_after_add(self) -> None:
        for spec in FIELDS:
            if spec.computed:
                continue
            w = self.input_widgets[spec.key]
            if spec.kind == "text":
                w.clear()
            else:
                if spec.key == "markup_percent":
                    w.setValue(60.0)
                else:
                    w.setValue(float(spec.default or 0.0))
        self._recompute_computed_fields()

    def _normalize_header(self, v: Any) -> str:
        return str(v or "").strip()

    def _to_float(self, v: Any) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(" ", "")
        if not s:
            return None
        s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    def import_from_excel(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Загрузить из Excel",
            "",
            "Excel (*.xlsx)",
        )
        if not file_path:
            return

        expected = [f.excel_header for f in FIELDS]
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            header = [self._normalize_header(ws.cell(row=1, column=i).value) for i in range(1, 5)]
            if header != expected:
                QMessageBox.warning(
                    self,
                    "Неверная структура Excel",
                    "Ожидались заголовки (A-D):\n"
                    f"- {expected[0]}\n- {expected[1]}\n- {expected[2]}\n- {expected[3]}\n\n"
                    "Проверьте первую строку файла.",
                )
                return

            added = 0
            for r_idx in range(2, ws.max_row + 1):
                name = self._normalize_header(ws.cell(row=r_idx, column=1).value)
                purchase = self._to_float(ws.cell(row=r_idx, column=2).value)
                markup = self._to_float(ws.cell(row=r_idx, column=3).value)
                retail = self._to_float(ws.cell(row=r_idx, column=4).value)

                if not name and purchase is None and markup is None and retail is None:
                    continue

                row: Row = {
                    "name": name,
                    "purchase_price": float(purchase or 0.0),
                    "markup_percent": float(markup if markup is not None else 60.0),
                }
                if retail is None:
                    row["retail_price"] = compute_retail(row)
                else:
                    row["retail_price"] = round(float(retail), 2)

                self.rows.append(row)
                self._append_row_to_table(row)
                added += 1
        except Exception as e:
            QMessageBox.critical(self, "Ошибка чтения Excel", str(e))
            return
        finally:
            try:
                wb.close()
            except Exception:
                pass

        QMessageBox.information(self, "Импорт завершён", f"Добавлено строк: {added}")

    def clear_rows(self) -> None:
        self.rows.clear()
        self.table.setRowCount(0)
        self._recompute_computed_fields()

    def delete_selected_rows(self) -> None:
        selection = self.table.selectionModel()
        if selection is None:
            return

        indexes = selection.selectedRows()
        if not indexes:
            QMessageBox.information(self, "Удаление", "Сначала выберите строку в таблице.")
            return

        row_indexes = sorted({idx.row() for idx in indexes}, reverse=True)
        for r in row_indexes:
            if 0 <= r < len(self.rows):
                self.rows.pop(r)
            self.table.removeRow(r)

    def export_excel(self) -> None:
        if not self.rows:
            QMessageBox.information(self, "Экспорт", "Список пуст — нечего сохранять.")
            return

        empty_names = [
            i + 1
            for i, row in enumerate(self.rows)
            if not str(row.get("name") or "").strip()
        ]
        if empty_names:
            shown = ", ".join(map(str, empty_names[:15]))
            suffix = "" if len(empty_names) <= 15 else "…"
            QMessageBox.warning(
                self,
                "Предупреждение",
                "Есть товары с пустым названием.\n"
                f"Строки: {shown}{suffix}\n\n"
                "Файл всё равно будет сохранён.",
            )

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"price_list_{ts}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет в Excel",
            default_name,
            "Excel (*.xlsx)",
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Price List"

        header_font = Font(bold=True)
        for col, spec in enumerate(FIELDS, start=1):
            cell = ws.cell(row=1, column=col, value=spec.excel_header)
            cell.font = header_font

        for r_idx, row in enumerate(self.rows, start=2):
            for c_idx, spec in enumerate(FIELDS, start=1):
                value = row.get(spec.key, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if spec.kind == "number" and spec.number_format:
                    cell.number_format = spec.number_format

        for c_idx, spec in enumerate(FIELDS, start=1):
            max_len = len(spec.excel_header)
            for row in self.rows:
                v = row.get(spec.key, "")
                if isinstance(v, float):
                    s = f"{v:.{spec.decimals}f}"
                else:
                    s = str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(
                max(10, max_len + 2), 48
            )

        try:
            wb.save(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения", str(e))
            return

        QMessageBox.information(self, "Готово", f"Сохранено:\n{file_path}")


def main() -> int:
    app = QApplication(sys.argv)
    w = PriceListApp()
    w.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())

